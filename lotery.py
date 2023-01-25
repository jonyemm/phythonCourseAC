import pycountry
import random

from openpyxl import Workbook
from openpyxl import load_workbook

with open("liste.txt") as liste:
    lineas = liste.readlines()

premios = []
ciudades = []
for n in range(len(lineas)):
  ciudades.append(list(pycountry.countries)[random.randint(0,len(pycountry.countries)-1)].name)


i = 0
totalPremios = 0
while i < len(lineas):
  monto = random.randint(50000,60000)
  premios.append(monto)
  totalPremios = sum(premios)
  i = i + 1

if(totalPremios < 1000000):
  restante = (1000000 - totalPremios)/len(lineas)
  i = 0
  while i < len(lineas):
    premios[i] = round(premios[i] + restante,2)
    i = i + 1


wb = Workbook()

filesheet = "./demosheet.xlsx"
wb.save(filesheet)

filesheet = "./demosheet.xlsx"

wb = load_workbook(filesheet)
sheet = wb.active

datos = [('nombre', 'premio', 'viaje')]
for row in datos:
 sheet.append(row)

i = 0
c = 2
while i < len(lineas):
  sheet['A'+str(c)] = lineas[i]
  sheet['B'+str(c)] = premios[i]
  sheet['C'+str(c)] = ciudades[i]
  i = i + 1
  c = c + 1

wb.save(filesheet)